#
# EXCEL_REPORT_GENERATOR.py - Generatore di report Excel multi-foglio
# Copyright 2025 TIM SPA
# Author Daniele Speziale
# Filename: Complex/excel_report_generator.py
# Created 29/09/25
# Description: Genera file Excel con più fogli da query multiple
#
import logging
import pandas as pd
from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from Complex.Core.database_manager import DatabaseManager


class ExcelReportGenerator:
    """Genera report Excel multi-foglio da query database"""

    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.db_manager = DatabaseManager(config)
        self.logger = logging.getLogger(__name__)

        # Directory output
        self.output_directory = Path(config.get('execution', {}).get('excel_output_directory', '../Reports'))
        self.output_directory.mkdir(parents=True, exist_ok=True)

        # Directory query SQL
        self.query_directory = Path(config.get('execution', {}).get('query_directory', '../Queries'))

    def execute_query(self, database_name: str, sql: str) -> pd.DataFrame:
        """Esegue una query e restituisce un DataFrame"""
        self.logger.info(f"EXCEL: Esecuzione query su [{database_name}]")

        db_config = self.db_manager.get_database_config(database_name)
        db_type = db_config.get('type')

        try:
            if db_type == 'oracle':
                return self._execute_oracle_query(database_name, sql)
            elif db_type == 'mssql':
                return self._execute_mssql_query(database_name, sql)
            elif db_type == 'sqlite':
                return self._execute_sqlite_query(database_name, sql)
            else:
                raise ValueError(f"Tipo database '{db_type}' non supportato")
        except Exception as e:
            self.logger.error(f"ERRORE: Query fallita su [{database_name}]: {e}")
            raise

    def _execute_oracle_query(self, database_name: str, sql: str) -> pd.DataFrame:
        """Esegue query Oracle"""
        with self.db_manager.get_oracle_connection(database_name) as conn:
            cursor = conn.cursor()
            cursor.execute(sql)

            columns = [col[0] for col in cursor.description]
            data = cursor.fetchall()

            return pd.DataFrame(data, columns=columns)

    def _execute_mssql_query(self, database_name: str, sql: str) -> pd.DataFrame:
        """Esegue query SQL Server"""
        with self.db_manager.get_connection(database_name) as conn:
            return pd.read_sql(sql, conn)

    def _execute_sqlite_query(self, database_name: str, sql: str) -> pd.DataFrame:
        """Esegue query SQLite"""
        with self.db_manager.get_connection(database_name) as conn:
            return pd.read_sql(sql, conn)

    def resolve_sql(self, sheet_config: Dict[str, Any]) -> str:
        """Risolve la query SQL da varie sorgenti"""
        sheet_name = sheet_config.get('name', 'unnamed')

        # SQL inline come stringa
        if 'sql' in sheet_config and isinstance(sheet_config['sql'], str):
            return sheet_config['sql']

        # SQL inline come array
        elif 'sql' in sheet_config and isinstance(sheet_config['sql'], list):
            sql = ' '.join(line.strip() for line in sheet_config['sql'])
            self.logger.info(f"EXCEL: SQL risolto da array per foglio [{sheet_name}]")
            return sql

        # SQL da file esterno
        elif 'sql_file' in sheet_config:
            sql_file = self.query_directory / sheet_config['sql_file']

            if not sql_file.exists():
                raise FileNotFoundError(f"File SQL non trovato: {sql_file}")

            with open(sql_file, 'r', encoding='utf-8') as f:
                sql = f.read()

            self.logger.info(f"EXCEL: SQL caricato da file [{sql_file.name}] per foglio [{sheet_name}]")
            return sql

        # SQL template con parametri
        elif 'sql_template' in sheet_config:
            template = sheet_config['sql_template']
            parameters = sheet_config.get('parameters', {})

            sql = template
            for param_name, param_value in parameters.items():
                placeholder = f"{{{param_name}}}"
                sql = sql.replace(placeholder, str(param_value))

            self.logger.info(f"EXCEL: Template SQL risolto per foglio [{sheet_name}]")
            return sql

        else:
            raise ValueError(f"Nessuna sorgente SQL trovata per foglio [{sheet_name}]")

    def apply_excel_formatting(self, worksheet, df: pd.DataFrame):
        """Applica formattazione professionale al foglio Excel"""

        # Stili header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Bordi
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Formatta header (prima riga)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Formatta dati con bordi e alternanza colori
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1), start=2):
            # Colore alternato per le righe
            if row_idx % 2 == 0:
                row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            else:
                row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            for cell in row:
                cell.border = thin_border
                cell.fill = row_fill
                cell.alignment = Alignment(vertical="center")

        # Auto-dimensiona colonne
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = min(length + 2, 50)  # Max 50 caratteri
            worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        # Freeze panes (blocca header)
        worksheet.freeze_panes = worksheet['A2']

    def set_document_properties(self, workbook: Workbook, report_config: Dict[str, Any]):
        """Imposta le proprietà del documento Excel"""
        properties = report_config.get('properties', {})

        if properties:
            # Proprietà standard Excel
            if 'title' in properties:
                workbook.properties.title = properties['title']

            if 'subject' in properties:
                workbook.properties.subject = properties['subject']

            if 'author' in properties:
                workbook.properties.creator = properties['author']

            if 'company' in properties:
                workbook.properties.company = properties['company']

            if 'category' in properties:
                workbook.properties.category = properties['category']

            if 'keywords' in properties:
                workbook.properties.keywords = properties['keywords']

            if 'comments' in properties:
                workbook.properties.description = properties['comments']

            if 'manager' in properties:
                workbook.properties.manager = properties['manager']

            # Timestamp automatico
            workbook.properties.created = datetime.now()
            workbook.properties.modified = datetime.now()

            self.logger.info(f"EXCEL: Proprietà documento impostate per report [{report_config.get('name')}]")

    def generate_report(self, report_config: Dict[str, Any]) -> str:
        """Genera un singolo report Excel con più fogli"""
        report_name = report_config.get('name', 'report')
        sheets = report_config.get('sheets', [])

        if not sheets:
            raise ValueError(f"Report [{report_name}] non ha fogli configurati")

        self.logger.info(f"EXCEL: Generazione report [{report_name}] con {len(sheets)} fogli")

        # Crea workbook
        wb = Workbook()
        wb.remove(wb.active)  # Rimuove foglio di default

        # Imposta proprietà documento
        self.set_document_properties(wb, report_config)

        results = {
            'report_name': report_name,
            'sheets_created': [],
            'errors': []
        }

        # Genera ogni foglio
        for sheet_config in sheets:
            sheet_name = sheet_config.get('name', 'Sheet')
            database_name = sheet_config.get('database')

            if not database_name:
                error_msg = f"Database non specificato per foglio [{sheet_name}]"
                self.logger.error(error_msg)
                results['errors'].append(error_msg)
                continue

            try:
                # Risolvi e esegui query
                sql = self.resolve_sql(sheet_config)
                df = self.execute_query(database_name, sql)

                # Crea foglio
                ws = wb.create_sheet(title=sheet_name[:31])  # Excel max 31 caratteri

                # Scrivi dati
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

                # Applica formattazione
                if len(df) > 0:
                    self.apply_excel_formatting(ws, df)

                self.logger.info(f"EXCEL: Foglio [{sheet_name}] creato con {len(df)} righe")
                results['sheets_created'].append({
                    'name': sheet_name,
                    'rows': len(df),
                    'columns': len(df.columns)
                })

            except Exception as e:
                error_msg = f"Errore creazione foglio [{sheet_name}]: {str(e)}"
                self.logger.error(error_msg)
                results['errors'].append(error_msg)

        # Salva file Excel
        if results['sheets_created']:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            #filename = f"{report_name}_{timestamp}.xlsx"
            filename = f"{report_name}.xlsx"
            filepath = self.output_directory / filename

            wb.save(filepath)
            self.logger.info(f"EXCEL: Report salvato in [{filepath}]")

            results['filepath'] = str(filepath)
            results['success'] = True
        else:
            results['success'] = False
            self.logger.error(f"EXCEL: Nessun foglio creato per report [{report_name}]")

        return results

    def generate_all_reports(self) -> List[Dict[str, Any]]:
        """Genera tutti i report configurati"""
        excel_reports = self.config.get('excel_reports', [])

        if not excel_reports:
            self.logger.warning("Nessun report Excel configurato")
            return []

        self.logger.info(f"EXCEL: Generazione di {len(excel_reports)} report")

        results = []
        for report_config in excel_reports:
            if not report_config.get('enabled', True):
                self.logger.info(f"EXCEL: Report [{report_config.get('name')}] disabilitato")
                continue

            try:
                result = self.generate_report(report_config)
                results.append(result)
            except Exception as e:
                self.logger.error(f"EXCEL: Errore generazione report: {e}")
                results.append({
                    'report_name': report_config.get('name', 'unknown'),
                    'success': False,
                    'errors': [str(e)]
                })

        return results